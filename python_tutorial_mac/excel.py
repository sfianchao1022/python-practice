import openpyxl as xl  # 把xl當作openpyxl的別名
from openpyxl.chart import BarChart, Reference  # import openpyxl(packages).chart(module).BarChart, Reference(class)

wb = xl.load_workbook("transactions.xlsx")  # excel存進工作表
sheet = wb['Sheet1']  # 選取工作表內的活頁簿Sheet1 一定要大寫
cell = sheet.cell(1, 1)  # 選取活頁簿中矩陣(1, 1)的格子
cell = sheet['a1']  # 也可以直接選定格子的編號
print(cell.value)  # 印出格子的值
print(sheet.max_row)  # 最多有4行
print(sheet.max_column)  # 最多有3列

for row in range(2, sheet.max_row + 1):  # 印 row 2 ~ 4 行
    cell = sheet.cell(row, 3)  # (2 ~ 4, 3)
    print(cell.value)
    corrected_price = cell.value * 0.9  # 操作格子內的值
    corrected_price_cell = sheet.cell(row, 4)  # 原本沒有column 4 所以新增column 4上的格子
    corrected_price_cell.value = corrected_price  # column 4上的格子給新值

# 選定活頁簿的範圍Reference (row 2 ~ 4, col = 4)
values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
chart = BarChart()  # 建立圖表物件
chart.add_data(values)  # chart給值
# sheet['e2'] = chart  # 圖表chart不能直接塞進格子裡 要用活頁簿的函式add_chart()
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')  # 另存成新檔
